# ============================================================
#  DAY 35 — Exporting pandas to Excel (Polished Multi-Sheet)
#  Author : Bhanu Pratap Singh
#  Date   : Day 35 of 84-Day Roadmap
# ============================================================

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# STEP 1 — Load raw data from input Excel
# ──────────────────────────────────────────────────────────────
# pd.read_excel reads the sheet into a DataFrame
df = pd.read_excel("day35_input.xlsx", sheet_name="Transactions")

# Convert Date column to actual datetime objects so we can group by month
df["Date"] = pd.to_datetime(df["Date"])

# ──────────────────────────────────────────────────────────────
# STEP 2 — Build derived DataFrames (transforms learned in Days 31-34)
# ──────────────────────────────────────────────────────────────

# Gain/Loss per transaction
df["Invested"]   = df["Shares"] * df["Buy_Price"]
df["Curr_Value"] = df["Shares"] * df["Current_Price"]
df["P&L"]        = df["Curr_Value"] - df["Invested"]
df["Return_%"]   = ((df["Curr_Value"] - df["Invested"]) / df["Invested"] * 100).round(2)

# Sector-level summary
sector_summary = (
    df.groupby("Sector")
      .agg(
          Total_Invested =("Invested",   "sum"),
          Current_Value  =("Curr_Value", "sum"),
          Total_PnL      =("P&L",        "sum"),
          Trades         =("Stock",      "count")
      )
      .reset_index()
)
sector_summary["Return_%"] = (
    (sector_summary["Total_PnL"] / sector_summary["Total_Invested"] * 100).round(2)
)

# Stock-level summary
stock_summary = (
    df.groupby("Stock")
      .agg(
          Total_Invested =("Invested",   "sum"),
          Current_Value  =("Curr_Value", "sum"),
          Total_PnL      =("P&L",        "sum"),
          Avg_Return     =("Return_%",   "mean")
      )
      .reset_index()
)
stock_summary["Avg_Return"] = stock_summary["Avg_Return"].round(2)

# Monthly P&L trend
df["Month"] = df["Date"].dt.to_period("M").astype(str)
monthly_pnl = (
    df.groupby("Month")
      .agg(Total_PnL=("P&L", "sum"), Trades=("Stock", "count"))
      .reset_index()
)

# ──────────────────────────────────────────────────────────────
# STEP 3 — Write all DataFrames to separate sheets using ExcelWriter
# NEW CONCEPT: pd.ExcelWriter  +  to_excel()  +  multi-sheet export
# ──────────────────────────────────────────────────────────────
output_path = "day35_output.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

    # Sheet 1 — Full transaction detail
    # startrow=1 leaves row 0 blank so we can add a title above the header
    df.to_excel(writer, sheet_name="Transactions", index=False, startrow=1)

    # Sheet 2 — Sector summary
    sector_summary.to_excel(writer, sheet_name="Sector_Summary", index=False, startrow=1)

    # Sheet 3 — Stock summary
    stock_summary.to_excel(writer, sheet_name="Stock_Summary", index=False, startrow=1)

    # Sheet 4 — Monthly trend
    monthly_pnl.to_excel(writer, sheet_name="Monthly_PnL", index=False, startrow=1)

# ExcelWriter context manager automatically saves + closes the file here

# ──────────────────────────────────────────────────────────────
# STEP 4 — Post-export formatting with openpyxl
# NEW CONCEPT: Load the saved file and add professional styling
# ──────────────────────────────────────────────────────────────

wb = load_workbook(output_path)

# ── Style helpers ────────────────────────────────────────────
def style_sheet(ws, title_text, header_color="1F4E79"):
    """Apply title + header formatting to any sheet."""

    # Title row (row 1)
    ws.insert_rows(1)          # push everything down by one row
    ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")
    title_cell = ws["A1"]
    title_cell.value     = title_text
    title_cell.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", start_color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row (now row 2 after insert)
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color=header_color)
    center       = Alignment(horizontal="center", vertical="center")
    thin         = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    for cell in ws[2]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin
    ws.row_dimensions[2].height = 18

    # Data rows
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4


# ── Apply to each sheet ──────────────────────────────────────
style_sheet(wb["Transactions"],  "📊 Portfolio Transactions — FY2026",  "2E75B6")
style_sheet(wb["Sector_Summary"],"📂 Sector-Wise Performance Summary",   "375623")
style_sheet(wb["Stock_Summary"], "📈 Stock-Level Performance Summary",   "843C0C")
style_sheet(wb["Monthly_PnL"],   "📅 Monthly P&L Trend",                 "7030A0")

# ──────────────────────────────────────────────────────────────
# STEP 5 — Conditional colour: green for profit, red for loss
# Applied to P&L columns on every relevant sheet
# ──────────────────────────────────────────────────────────────
green_font = Font(color="375623", bold=True, name="Arial")
red_font   = Font(color="C00000", bold=True, name="Arial")

for sheet_name in ["Transactions", "Sector_Summary", "Stock_Summary"]:
    ws = wb[sheet_name]
    # Find column with "P&L" or "PnL" in header (row 2)
    pnl_col = None
    for cell in ws[2]:
        if cell.value and ("P&L" in str(cell.value) or "PnL" in str(cell.value)):
            pnl_col = cell.column
            break
    if pnl_col:
        for row in ws.iter_rows(min_row=3, min_col=pnl_col, max_col=pnl_col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.font = green_font if cell.value >= 0 else red_font

wb.save(output_path)
print(f"✅  Saved → {output_path}")
print(f"    Sheets : {wb.sheetnames}")
print(f"    Total rows (Transactions): {len(df)}")
